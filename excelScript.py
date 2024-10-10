import openpyxl
import openpyxl.styles
import database


# Метод для форматирования границ таблицы
def set_border(ws, cell_range, need_to_thick, need_to_thick_up, need_to_thick_down):
    thin = openpyxl.styles.Side(border_style="thin", color="000000")
    thick = openpyxl.styles.Side(border_style="thick", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            if cell == row[len(row) - 1]:
                cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thick, bottom=thin)
                cell.fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            else:
                cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
        if len(need_to_thick) != 0:
            if row == ws[need_to_thick[0]]:
                for cell in row:
                    cell.border = openpyxl.styles.Border(top=thick, left=thick, right=thick, bottom=thick)
                need_to_thick.pop(0)
        if len(need_to_thick_up) != 0:
            if row == ws[need_to_thick_up[0]]:
                for cell in row:
                    if cell == row[len(row) - 1]:
                        cell.border = openpyxl.styles.Border(top=thick, left=thin, right=thick, bottom=thin)
                    else:
                        cell.border = openpyxl.styles.Border(top=thick, left=thin, right=thin, bottom=thin)
                need_to_thick_up.pop(0)
        if len(need_to_thick_down) != 0:
            if row == ws[need_to_thick_down[0]]:
                for cell in row:
                    if cell == row[len(row) - 1]:
                        cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thick, bottom=thick)
                    else:
                        cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thick)
                need_to_thick_down.pop(0)


# Метод для формирования Excel таблицы с составом апелляционной комиссии
def make_commission_excel():
    # Получаем состав комиссии из базы данных
    data = database.Database.get_commission_members()
    
    # Создаем книгу и заполняем ее
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Состав комиссии'

    # Формирование шапки таблицы
    header_data = ["No", "Ф.И.О Эксперта", "Должность"]
    for col, header in enumerate(header_data, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Заполнение данных
    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Установка ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 100

    # Установка границ таблицы
    set_border(ws, f'A1:C{len(data) + 1}', [], [], [])

    # Сохранение файла
    wb.save('Состав_апелляционной_комиссии.xlsx')
    print("Файл успешно сохранен")


make_commission_excel()
